import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, GradientFill, Color
import pandas as pd
   
export_val = pd.read_csv('D:\python\Trialink\Trialink_prjct\Trialink\\test_export.csv')
wb = Workbook()
ws = wb.active
font = Font(color='FF0000')
ws[export_val].font = font
wb.save('result_test.xlsx')